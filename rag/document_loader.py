"""Загрузка и парсинг документов (PDF, DOCX, TXT, MD, CSV, XLSX)."""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class DocumentLoadError(Exception):
    """Ошибка при загрузке документа."""


class DocumentLoader:
    """Загружает текст из файлов различных форматов."""

    def load(self, file_path: str) -> str:
        """Загрузить текст из файла.

        Args:
            file_path: Путь к файлу.

        Returns:
            Извлечённый текст.

        Raises:
            DocumentLoadError: Если файл не удалось прочитать.
        """
        path = Path(file_path)
        if not path.exists():
            raise DocumentLoadError(f"Файл не найден: {file_path}")

        ext = path.suffix.lower()
        try:
            if ext == ".pdf":
                return self._load_pdf(path)
            elif ext == ".docx":
                return self._load_docx(path)
            elif ext in (".txt", ".md"):
                return self._load_text(path)
            elif ext == ".csv":
                return self._load_csv(path)
            elif ext == ".xlsx":
                return self._load_xlsx(path)
            else:
                raise DocumentLoadError(f"Неподдерживаемый формат файла: {ext}")
        except DocumentLoadError:
            raise
        except Exception as e:
            raise DocumentLoadError(f"Ошибка при чтении файла {path.name}: {e}") from e

    def _load_pdf(self, path: Path) -> str:
        """Извлечь текст из PDF."""
        from PyPDF2 import PdfReader

        reader = PdfReader(str(path))
        pages: list[str] = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        if not pages:
            raise DocumentLoadError("PDF не содержит извлекаемого текста")
        return "\n\n".join(pages)

    def _load_docx(self, path: Path) -> str:
        """Извлечь текст из DOCX."""
        from docx import Document

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        if not paragraphs:
            raise DocumentLoadError("DOCX не содержит текста")
        return "\n\n".join(paragraphs)

    def _load_text(self, path: Path) -> str:
        """Прочитать текстовый файл с автоопределением кодировки."""
        import chardet

        raw = path.read_bytes()
        if not raw:
            raise DocumentLoadError("Файл пуст")
        detected = chardet.detect(raw)
        encoding = detected.get("encoding", "utf-8") or "utf-8"
        return raw.decode(encoding)

    def _load_csv(self, path: Path) -> str:
        """Загрузить CSV как текст с заголовками."""
        import csv
        import chardet

        raw = path.read_bytes()
        if not raw:
            raise DocumentLoadError("CSV файл пуст")
        detected = chardet.detect(raw)
        encoding = detected.get("encoding", "utf-8") or "utf-8"
        text = raw.decode(encoding)

        rows: list[str] = []
        reader = csv.reader(text.splitlines())
        headers: list[str] = []
        for i, row in enumerate(reader):
            if i == 0:
                headers = row
                continue
            if headers:
                parts = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
                rows.append(" | ".join(parts))
            else:
                rows.append(" | ".join(row))
        if not rows:
            raise DocumentLoadError("CSV не содержит данных")
        return "\n".join(rows)

    def _load_xlsx(self, path: Path) -> str:
        """Загрузить Excel как текст."""
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        all_text: list[str] = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            all_text.append(f"Лист: {sheet}")
            headers = [str(c) if c is not None else "" for c in rows[0]]

            for row in rows[1:]:
                values = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in values):
                    if headers:
                        parts = [f"{h}: {v}" for h, v in zip(headers, values) if v.strip()]
                        all_text.append(" | ".join(parts))
                    else:
                        all_text.append(" | ".join(values))

        wb.close()
        if not all_text:
            raise DocumentLoadError("Excel файл не содержит данных")
        return "\n".join(all_text)

    def load_url(self, url: str) -> str:
        """Загрузить текст с веб-страницы.

        Args:
            url: URL страницы.

        Returns:
            Извлечённый текст.
        """
        import httpx
        from bs4 import BeautifulSoup

        try:
            response = httpx.get(url, timeout=30.0, follow_redirects=True)
            response.raise_for_status()
        except Exception as e:
            raise DocumentLoadError(f"Не удалось загрузить URL: {e}") from e

        soup = BeautifulSoup(response.text, "html.parser")

        # Удаляем скрипты и стили
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()

        text = soup.get_text(separator="\n", strip=True)
        # Убираем пустые строки
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        result = "\n".join(lines)

        if len(result) < 50:
            raise DocumentLoadError("Страница не содержит достаточно текста")
        return result
